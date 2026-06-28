import json
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ==========================================
# Load configuration
# ==========================================

CONFIG_PATH = Path("config/config.json")

with open(CONFIG_PATH, "r") as file:
    config = json.load(file)

RAW_FILE = config["dirty_output_file"]
CLEAN_FILE = config["clean_output_file"]
REPORT_FILE = config["quality_report"]

# ==========================================
# Load datasets
# ==========================================

print("\nLoading datasets...")

raw_df = pd.read_csv(RAW_FILE)
clean_df = pd.read_csv(CLEAN_FILE)

# ==========================================
# Calculate Metrics
# ==========================================

original_records = len(raw_df)
clean_records = len(clean_df)

records_removed = original_records - clean_records

duplicates = raw_df.duplicated(
    subset=["CustomerName", "Email"]
).sum()

invalid_emails = (
    ~raw_df["Email"]
    .fillna("")
    .str.match(r"^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$")
).sum()

missing_values = raw_df.isna().sum().sum()

quality_score = round(
    (clean_records / original_records) * 100,
    2
)

# ==========================================
# Create Workbook
# ==========================================

wb = Workbook()

header_fill = PatternFill(
    fill_type="solid",
    fgColor="1F4E78"
)

header_font = Font(
    color="FFFFFF",
    bold=True
)

title_font = Font(
    size=18,
    bold=True
)

# ==========================================
# Sheet 1 - Executive Summary
# ==========================================

ws = wb.active
ws.title = "Executive Summary"

ws["A1"] = "Enterprise Data Quality Report"
ws["A1"].font = title_font

metrics = [
    ("Original Records", original_records),
    ("Clean Records", clean_records),
    ("Records Removed", records_removed),
    ("Duplicate Records", duplicates),
    ("Invalid Emails", invalid_emails),
    ("Missing Values", missing_values),
    ("Quality Score (%)", quality_score)
]

row = 3

for metric, value in metrics:

    ws[f"A{row}"] = metric
    ws[f"B{row}"] = value

    ws[f"A{row}"].fill = header_fill
    ws[f"A{row}"].font = header_font
    ws[f"A{row}"].alignment = Alignment(horizontal="center")

    row += 1

# ==========================================
# Sheet 2 - Dataset Preview
# ==========================================

preview = wb.create_sheet("Dataset Preview")

for col, column in enumerate(clean_df.columns, start=1):

    cell = preview.cell(row=1, column=col)

    cell.value = column
    cell.fill = header_fill
    cell.font = header_font

for r, values in enumerate(
    clean_df.head(100).values,
    start=2
):

    for c, value in enumerate(values, start=1):

        preview.cell(
            row=r,
            column=c
        ).value = value

# ==========================================
# Sheet 3 - Statistics
# ==========================================

stats = wb.create_sheet("Statistics")

stats["A1"] = "Data Quality Statistics"
stats["A1"].font = title_font

statistics = {
    "Duplicate %":
        round((duplicates / original_records) * 100, 2),

    "Invalid Email %":
        round((invalid_emails / original_records) * 100, 2),

    "Missing Values %":
        round((missing_values / (original_records * len(raw_df.columns))) * 100, 2),

    "Retention %":
        round((clean_records / original_records) * 100, 2)
}

row = 3

for metric, value in statistics.items():

    stats[f"A{row}"] = metric
    stats[f"B{row}"] = value

    stats[f"A{row}"].fill = header_fill
    stats[f"A{row}"].font = header_font

    row += 1

# ==========================================
# Save Report
# ==========================================

Path("output").mkdir(exist_ok=True)

wb.save(REPORT_FILE)

print("\nExecutive Excel report created successfully.")
print(f"Report: {REPORT_FILE}")